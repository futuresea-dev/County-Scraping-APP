from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import time
import random
import undetected_chromedriver.v2 as uc
from os.path import dirname, abspath
import os
import csv
import pandas as pd
from multiprocessing import Pool, freeze_support
import requests
import io
from PIL import Image
import pytesseract
from wand.image import Image as wi
import time
from decimal import Decimal
from openpyxl import load_workbook

write_header = True
BASE = os.path.dirname(abspath(__file__))


def get_proxies():
    with open(os.path.join(BASE, "proxy.txt")) as f:
        content = f.readlines()
    # you may also want to remove whitespace characters like `\n` at the end of each line
    PROXIES = [x.strip() for x in content]

    return PROXIES


def chunks(l, n):
    """Yield successive n-sized chunks from l."""
    for i in range(0, len(l), n):
        yield l[i:i + n]


def proxy_driver(PROXIES):
    co = uc.ChromeOptions()

    co.add_argument("--disable-extensions")
    co.add_argument("--disable-popup-blocking")
    co.add_argument("--profile-directory=Default")
    co.add_argument("--disable-plugins-discovery")
    co.add_argument("--incognito")
    co.add_argument("--headless")
    co.add_argument('--no-sandbox')
    co.add_argument("--disable-setuid-sandbox")
    co.add_argument("user_agent=DN")
    co.add_argument("--start-maximized")
    pxy = random.choice(PROXIES)
    # pxy = "2.56.46.10:8800"

    co.add_argument('--proxy-server=%s' % pxy)

    # driver = uc.Chrome(options=co)
    driver = uc.Chrome(options=co, executable_path="chromedriver.exe")
    driver.delete_all_cookies()

    return driver


def scraping_tax(driver, p_id):
    global write_header
    output = []
    takeout_url = "https://travis.go2gov.net/cart/responsive/search/displayQuickSearch.do?"
    driver.get(takeout_url)
    delay = 5
    try:
        WebDriverWait(driver, delay).until(
            EC.presence_of_element_located((By.ID, 'qsfInput')))
        id_input = driver.find_element_by_id("qsfInput")
        id_input.clear()
        id_input.send_keys(p_id)

        search_button = driver.find_element_by_id("qsfButtonSearch")
        driver.execute_script("arguments[0].click();", search_button)
        time.sleep(1)

        # get tbody
        search_result = driver.find_element_by_tag_name("tbody")
        search_trs = search_result.find_elements_by_tag_name("tr")
        tax_urls = []
        for search_tr in search_trs:
            each_tax = search_tr.find_element_by_xpath("//td[1]/div[1]/a")
            # tax_controller_map = each_tax.text
            tax_controller_url = each_tax.get_attribute("href")
            tax_urls.append(tax_controller_url)

        for tax_url in tax_urls:
            out_list = ["", "Travis"]
            driver.get(tax_url)
            ownership_infos = driver.find_elements_by_css_selector(".three.columns")
            for idx, owner in enumerate(ownership_infos):
                if idx == 0:
                    pass
                elif idx == 1:
                    tax_controller_map = owner.find_element_by_tag_name("a").text
                    out_list.append(tax_controller_map)
                    out_list.append("")
                    out_list.append("")
                elif idx == 2:
                    owner_name = owner.text.replace("Owner Name\n", "")
                    out_list.append(owner_name)
                    out_list.append("")
                elif idx == 3:
                    address = owner.text.replace("Mailing Address\n", "")
                    mailings = address.split("\n")[::-1]
                    out_list.append(mailings[1])
                    out_list.append("")
                    out_list.append("")
                    mailing_info = mailings[0].split(",")
                    out_list.append(mailing_info[0].replace(",", ""))
                    out_list.append(mailing_info[1].split(" ")[1].replace(
                        ",", ""))
                    mailing_zip = mailing_info[1].split(" ")[2].split("-")
                    out_list.append(mailing_zip[0])
                    try:
                        out_list.append(mailing_zip[1])
                    except:
                        out_list.append("")
                    out_list = out_list + ["", "", "", "", "", "", "", "", "", "", "", ""]

            bill_url = driver.find_element_by_xpath('//*[@id="nav-buttons"]/div/div/div[8]/a').get_attribute("href")
            get_proxy = get_proxies()
            pxy = random.choice(get_proxy)
            proxies = {
                'http': 'http://' + pxy,
                'https': 'http://' + pxy,
            }
            response = requests.get(bill_url, proxies=proxies)
            pdf_file = time.strftime("%Y%m%d%H%M%S") + ".pdf"
            with open(pdf_file, 'wb') as f:
                f.write(response.content)
            try:
                pdfFile = wi(filename=pdf_file, resolution=200)
                image = pdfFile.convert('jpeg')

                imageBlobs = []

                for img in image.sequence:
                    imgPage = wi(image=img)
                    imageBlobs.append(imgPage.make_blob('jpeg'))

                extract = []

                for imgBlob in imageBlobs:
                    image = Image.open(io.BytesIO(imgBlob))
                    text = pytesseract.image_to_string(image, lang='eng')
                    extract.append(text)

                result = extract[0]
                lists = result.split("\n")
                years_tax = []
                tax_first = True
                for faxs in lists:
                    fax_list = faxs.split(" ")
                    if len(fax_list) == 5:
                        if fax_list[0].isnumeric() and int(fax_list[0]) > 1900 and int(fax_list[0]) < 2021:
                            years_tax.append(fax_list[4])
                if len(years_tax) > 0:
                    taxes = [""] * (31 - len(years_tax))
                    tax_res = taxes + years_tax[::-1]
                    out_list = out_list + tax_res
                else:
                    out_list = out_list + [""] * 31
                sum = 0
                for each_year in years_tax:
                    sum = sum + Decimal(each_year.replace(",", ""))
                out_list.append(f"{sum:,}")
                os.remove(pdf_file)
            except Exception as e:
                print(str(e))
            output.append(out_list)


    except TimeoutException:
        print("Page Loading took too much time!")
    return output


def main(Parcel_id):
    ALL_PROXIES = get_proxies()
    # --- YOU ONLY NEED TO CARE FROM THIS LINE ---
    # creating new driver to use proxy
    pd = proxy_driver(ALL_PROXIES)
    sc_result = []
    try:
        sc_result = scraping_tax(pd, Parcel_id)
        # new = ALL_PROXIES.pop()
        pd.close()

    except:
        print("scraping error")
        time.sleep(1)
        pd.close()
        pass
    return sc_result


if __name__ == "__main__":
    df = pd.ExcelFile('Parcel IDs_3.xlsx').parse('Sheet1')  # you could add index_col=0 if there's an index
    x = [df['Parcel ID']]
    p_ids = x[0]
    # p_ids.pop(0)
    result = list(chunks(p_ids, 5))
    freeze_support()
    for res in result:
        lis = list(res)
        with Pool(processes=5) as pool:
            output = pool.map(main, lis)
            try:
                wb = load_workbook("output.xlsx")
                # Select First Worksheet
                ws = wb.worksheets[0]

                # Append 2 new Rows - Columns A - D
                for row_data in output:
                    # Append Row Values
                    ws.append(row_data)

                wb.save("output.xlsx")
            except:
                print("write xlsx error")
    # df = pd.ExcelFile('Parcel IDs_2.xlsx').parse('Sheet1')  # you could add index_col=0 if there's an index
    # x = [df['geo_id']]
    # p_ids = x[0]
    # p_ids.pop(0)
    # for i, p_id in enumerate(p_ids):
    #     output = main(p_id)
    #     try:
    #         wb = load_workbook("output.xlsx")
    #         # Select First Worksheet
    #         ws = wb.worksheets[0]
    #
    #         # Append 2 new Rows - Columns A - D
    #         for row_data in output:
    #             # Append Row Values
    #             ws.append(row_data)
    #
    #         wb.save("output.xlsx")
    #     except:
    #         print("write xlsx error")
    print("done!")
